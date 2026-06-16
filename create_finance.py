import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule
import os

OUTPUT = r"C:\Users\OperSystem\Documents\webtienda\Personal Finance Manager.xlsx"

wb = openpyxl.Workbook()

# Colors
DARK = "1a1a2e"
ACCENT = "e94560"
ACCENT2 = "0f3460"
WHITE = "ffffff"
LIGHT = "f0f0f0"
GREEN = "2ecc71"
RED_ = "e74c3c"
GRAY = "888888"

hdr_font = Font(name="Segoe UI", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=ACCENT)
hdr_fill2 = PatternFill("solid", fgColor=ACCENT2)
title_font = Font(name="Segoe UI", bold=True, color=DARK, size=16)
sub_font = Font(name="Segoe UI", bold=True, color=DARK, size=12)
body_font = Font(name="Segoe UI", size=10, color="333333")
money_fmt = '#,##0.00'
pct_fmt = '0%'
thin = Side(style='thin', color='cccccc')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
light_fill = PatternFill("solid", fgColor="f9f9f9")

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

def style_body(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.border = border
        if row % 2 == 0:
            cell.fill = light_fill

# ========== SHEET 1: DASHBOARD ==========
ws = wb.active
ws.title = "Dashboard"
ws.sheet_properties.tabColor = ACCENT
ws.merge_cells('A1:F1')
ws['A1'] = "Personal Finance Manager"
ws['A1'].font = Font(name="Segoe UI", bold=True, color=DARK, size=22)
ws['A2'] = "Financial Dashboard"
ws['A2'].font = Font(name="Segoe UI", color=GRAY, size=12)

# Summary cards
for r, label, val, fill in [
    (4, "Total Income", "=SUM(Transactions!G:G)", GREEN),
    (5, "Total Expenses", "=SUM(Transactions!H:H)", RED_),
    (6, "Net Savings", "=B4-B5", ACCENT),
    (7, "Monthly Budget", 5000, ACCENT2),
]:
    ws.cell(row=r, column=1, value=label).font = Font(name="Segoe UI", bold=True, size=10, color=WHITE)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=fill)
    ws.cell(row=r, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=3, value=val).font = Font(name="Segoe UI", bold=True, size=14, color=DARK)
    ws.cell(row=r, column=3).number_format = money_fmt

ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 22

# Charts
ws.cell(row=9, column=1, value="Expenses by Category").font = sub_font
ws.cell(row=9, column=4, value="Income vs Expenses").font = sub_font

# Budget vs actual
ws.cell(row=14, column=1, value="Budget vs Actual").font = sub_font
headers = ["Category", "Budget", "Actual", "Remaining", "% Used"]
for i, h in enumerate(headers, 1):
    ws.cell(row=15, column=i, value=h)
style_header(ws, 15, 5)
cats = ["Housing", "Food", "Transport", "Entertainment", "Utilities", "Insurance", "Savings", "Other"]
budgets = [1200, 600, 400, 200, 300, 250, 500, 200]
for idx, (c, b) in enumerate(zip(cats, budgets), 16):
    ws.cell(row=idx, column=1, value=c)
    ws.cell(row=idx, column=2, value=b)
    ws.cell(row=idx, column=2).number_format = money_fmt
    ws.cell(row=idx, column=3, value=f'=SUMIF(Categories!A:A,A{idx},Categories!B:B)')
    ws.cell(row=idx, column=3).number_format = money_fmt
    ws.cell(row=idx, column=4, value=f'=B{idx}-C{idx}')
    ws.cell(row=idx, column=4).number_format = money_fmt
    ws.cell(row=idx, column=5, value=f'=IF(B{idx}=0,0,C{idx}/B{idx})')
    ws.cell(row=idx, column=5).number_format = pct_fmt
    style_body(ws, idx, 5)

# Pie chart
pie = PieChart()
pie.title = "Expenses by Category"
pie.style = 10
pie.width = 16
pie.height = 12
data = Reference(ws, min_col=3, min_row=15, max_row=23)
cats_ref = Reference(ws, min_col=1, min_row=16, max_row=23)
pie.add_data(data, titles_from_data=True)
pie.set_categories(cats_ref)
pie.dataLabels = openpyxl.chart.label.DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showCatName = True
ws.add_chart(pie, "D9")

# ========== SHEET 2: TRANSACTIONS ==========
ws2 = wb.create_sheet("Transactions")
ws2.sheet_properties.tabColor = ACCENT2
ws2.merge_cells('A1:H1')
ws2['A1'] = "Income & Expense Log"
ws2['A1'].font = Font(name="Segoe UI", bold=True, color=DARK, size=18)

headers2 = ["Date", "Description", "Category", "Type", "Payment Method", "Status", "Income", "Expense"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, 8)

sample = [
    ["1/1/2026", "Salary - January", "Salary", "Income", "Bank Transfer", "Cleared", 5000, 0],
    ["1/2/2026", "Rent Payment", "Housing", "Expense", "Bank Transfer", "Cleared", 0, 1200],
    ["1/3/2026", "Groceries - Walmart", "Food", "Expense", "Credit Card", "Cleared", 0, 350],
    ["1/5/2026", "Gas Station", "Transport", "Expense", "Cash", "Pending", 0, 80],
    ["1/7/2026", "Netflix Subscription", "Entertainment", "Expense", "Credit Card", "Cleared", 0, 15.99],
    ["1/10/2026", "Freelance Project", "Freelance", "Income", "PayPal", "Cleared", 1200, 0],
]
for idx, row in enumerate(sample, 4):
    for c, v in enumerate(row, 1):
        ws2.cell(row=idx, column=c, value=v)
        if c == 7 or c == 8:
            ws2.cell(row=idx, column=c).number_format = money_fmt
    style_body(ws2, idx, 8)

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 30
ws2.column_dimensions['C'].width = 16
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 16
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 16
ws2.column_dimensions['H'].width = 16
ws2.freeze_panes = 'A4'

# Data validation for Type
from openpyxl.worksheet.datavalidation import DataValidation
dv = DataValidation(type="list", formula1='"Income,Expense"', allow_blank=True)
dv.error = "Select Income or Expense"
dv.errorTitle = "Invalid Type"
ws2.add_data_validation(dv)
dv.add('D4:D1000')

# ========== SHEET 3: CATEGORIES ==========
ws3 = wb.create_sheet("Categories")
ws3.sheet_properties.tabColor = GREEN
ws3['A1'] = "Category Master"
ws3['A1'].font = Font(name="Segoe UI", bold=True, color=DARK, size=14)
for i, h in enumerate(["Category", "Budget", "Type"], 1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, 3)
cat_data = [
    ["Salary", 5000, "Income"], ["Freelance", 2000, "Income"],
    ["Investments", 1000, "Income"],
    ["Housing", 1200, "Expense"], ["Food", 600, "Expense"],
    ["Transport", 400, "Expense"], ["Entertainment", 200, "Expense"],
    ["Utilities", 300, "Expense"], ["Insurance", 250, "Expense"],
    ["Savings", 500, "Expense"], ["Healthcare", 200, "Expense"],
    ["Education", 150, "Expense"], ["Shopping", 300, "Expense"],
    ["Other", 200, "Expense"],
]
for idx, row in enumerate(cat_data, 4):
    for c, v in enumerate(row, 1):
        ws3.cell(row=idx, column=c, value=v)
        if c == 2:
            ws3.cell(row=idx, column=c).number_format = money_fmt
    style_body(ws3, idx, 3)
ws3.column_dimensions['A'].width = 18
ws3.column_dimensions['B'].width = 14
ws3.column_dimensions['C'].width = 12

# ========== SHEET 4: MONTHLY SUMMARY ==========
ws4 = wb.create_sheet("Monthly Summary")
ws4.sheet_properties.tabColor = "3498db"
ws4['A1'] = "Monthly Reports"
ws4['A1'].font = Font(name="Segoe UI", bold=True, color=DARK, size=14)
for i, h in enumerate(["Month", "Income", "Expenses", "Net", "Savings Rate"], 1):
    ws4.cell(row=3, column=i, value=h)
style_header(ws4, 3, 5)
months = ["January", "February", "March", "April", "May", "June",
          "July", "August", "September", "October", "November", "December"]
for idx, m in enumerate(months, 4):
    ws4.cell(row=idx, column=1, value=m)
    ws4.cell(row=idx, column=2, value=f'=SUMIFS(Transactions!G:G,Transactions!A:A,">="&DATE(2026,{idx-3},1),Transactions!A:A,"<="&DATE(2026,{idx-3},EOMONTH(DATE(2026,{idx-3},1),0)))')
    ws4.cell(row=idx, column=2).number_format = money_fmt
    ws4.cell(row=idx, column=3, value=f'=SUMIFS(Transactions!H:H,Transactions!A:A,">="&DATE(2026,{idx-3},1),Transactions!A:A,"<="&DATE(2026,{idx-3},EOMONTH(DATE(2026,{idx-3},1),0)))')
    ws4.cell(row=idx, column=3).number_format = money_fmt
    ws4.cell(row=idx, column=4, value=f'=B{idx}-C{idx}')
    ws4.cell(row=idx, column=4).number_format = money_fmt
    ws4.cell(row=idx, column=5, value=f'=IF(B{idx}=0,0,D{idx}/B{idx})')
    ws4.cell(row=idx, column=5).number_format = pct_fmt
    style_body(ws4, idx, 5)

ws4.column_dimensions['A'].width = 16
ws4.column_dimensions['B'].width = 16
ws4.column_dimensions['C'].width = 16
ws4.column_dimensions['D'].width = 16
ws4.column_dimensions['E'].width = 16

# Bar chart
bar = BarChart()
bar.title = "Income vs Expenses"
bar.style = 10
bar.width = 22
bar.height = 14
data = Reference(ws4, min_col=2, max_col=3, min_row=3, max_row=15)
cats = Reference(ws4, min_col=1, min_row=4, max_row=15)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.y_axis.title = "Amount ($)"
ws4.add_chart(bar, "G3")

# ========== SHEET 5: SAVINGS GOALS ==========
ws5 = wb.create_sheet("Savings Goals")
ws5.sheet_properties.tabColor = "2ecc71"
ws5['A1'] = "Savings Goals Tracker"
ws5['A1'].font = Font(name="Segoe UI", bold=True, color=DARK, size=14)
for i, h in enumerate(["Goal", "Target", "Saved", "Remaining", "% Complete", "Deadline"], 1):
    ws5.cell(row=3, column=i, value=h)
style_header(ws5, 3, 6)
goals = [
    ["Emergency Fund", 10000, 4500, "Dec 2026"],
    ["New Car", 25000, 5000, "Jun 2027"],
    ["Vacation", 5000, 2000, "Aug 2026"],
    ["Education", 15000, 3000, "Jan 2028"],
]
for idx, row in enumerate(goals, 4):
    ws5.cell(row=idx, column=1, value=row[0])
    ws5.cell(row=idx, column=2, value=row[1])
    ws5.cell(row=idx, column=2).number_format = money_fmt
    ws5.cell(row=idx, column=3, value=row[2])
    ws5.cell(row=idx, column=3).number_format = money_fmt
    ws5.cell(row=idx, column=4, value=f'=B{idx}-C{idx}')
    ws5.cell(row=idx, column=4).number_format = money_fmt
    ws5.cell(row=idx, column=5, value=f'=IF(B{idx}=0,0,C{idx}/B{idx})')
    ws5.cell(row=idx, column=5).number_format = pct_fmt
    ws5.cell(row=idx, column=6, value=row[3])
    style_body(ws5, idx, 6)
    # Data bars for % complete
    ws5.conditional_formatting.add(f'E{idx}', DataBarRule(start_type='min', end_type='max', color=ACCENT))

ws5.column_dimensions['A'].width = 20
ws5.column_dimensions['B'].width = 14
ws5.column_dimensions['C'].width = 14
ws5.column_dimensions['D'].width = 14
ws5.column_dimensions['E'].width = 14
ws5.column_dimensions['F'].width = 14

# ========== SHEET 6: HOW TO USE ==========
ws6 = wb.create_sheet("How to Use")
ws6.sheet_properties.tabColor = GRAY
ws6['A1'] = "How to Use This Template"
ws6['A1'].font = Font(name="Segoe UI", bold=True, color=DARK, size=18)
instructions = [
    "📌 PERSONAL FINANCE MANAGER - USER GUIDE",
    "",
    "1. TRANSACTIONS SHEET",
    "   - Log all your income and expenses here",
    "   - Select Type: Income or Expense",
    "   - Income goes in column G, Expense in column H",
    "   - Use the dropdown for transaction type",
    "",
    "2. CATEGORIES SHEET",
    "   - Edit/add your custom categories here",
    "   - Set monthly budget for each expense category",
    "   - Categories drive the reports and dashboard",
    "",
    "3. DASHBOARD",
    "   - Auto-updates with your data",
    "   - See pie chart of expenses by category",
    "   - Track budget vs actual spending",
    "   - Monitor your net savings",
    "",
    "4. MONTHLY SUMMARY",
    "   - View income/expense trends month by month",
    "   - Bar chart shows visual comparison",
    "   - Track your savings rate over time",
    "",
    "5. SAVINGS GOALS",
    "   - Set financial goals with target amounts",
    "   - Track progress with visual data bars",
    "   - See percentage complete automatically",
    "",
    "💡 Pro Tips:",
    "   - Filter transactions by category or type",
    "   - Use PivotTables for deeper analysis",
    "   - Save a backup monthly",
    "   - Customize categories to your lifestyle",
    "",
    "Created by SystemsAcevedo © 2026",
    "Contact: acevedomoradavid433@gmail.com",
]
for idx, line in enumerate(instructions, 3):
    ws6.cell(row=idx, column=1, value=line)
    if line.startswith("📌") or line.startswith("💡"):
        ws6.cell(row=idx, column=1).font = Font(name="Segoe UI", bold=True, color=ACCENT, size=12)
    elif line.startswith("Created"):
        ws6.cell(row=idx, column=1).font = Font(name="Segoe UI", italic=True, color=GRAY, size=10)
    elif line.strip().isdigit() or (line.strip() and line.strip()[0].isdigit()):
        ws6.cell(row=idx, column=1).font = Font(name="Segoe UI", bold=True, color=DARK, size=10)
ws6.column_dimensions['A'].width = 60

# Save
wb.save(OUTPUT)
print(f"CREATED: {OUTPUT}")
print(f"Size: {os.path.getsize(OUTPUT)/1024:.1f} KB")
